import openpyxl

def create_excel_report(robot_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Общая сводка'

    for item in robot_data:
        model = item['model']
        version = item['version']
        count = item['total_count']

        if not model:
            continue

        if model not in wb.sheetnames:
            ws = wb.create_sheet(title=model)
            ws['A1'] = 'Модель'
            ws['B1'] = 'Версия'
            ws['C1'] = 'Количество за неделю'

        ws = wb[model]
        ws.append([model, version, count])

    if 'Общая сводка' in wb.sheetnames:
        wb.remove(wb['Общая сводка'])

    wb.save('robot_summary.xlsx')
